import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def read_file(file):
    if file.endswith('.csv'):
        return pd.read_csv(file, header=None)
    elif file.endswith('.xlsx'):
        return pd.read_excel(file, header=None)
    else:
        raise ValueError("Unsupported file format. Please provide a .csv or .xlsx file.")

def compare_files(file1, file2, file3):
    try:
        # Load the data from the files
        df1 = read_file(file1)
        df2 = read_file(file2)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        print(f"Please check if the files {file1} and {file2} exist.")
        return

    # Create a new workbook for the result
    wb = Workbook()
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Compare each cell
    for i in range(df1.shape[0]):
        for j in range(df1.shape[1]):
            cell_value1 = df1.iloc[i, j] if i < len(df1) and j < len(df1.columns) else None
            cell_value2 = df2.iloc[i, j] if i < len(df2) and j < len(df2.columns) else None

            if cell_value1 == cell_value2:
                ws.cell(row=i+1, column=j+1).value = f"{file1}: {cell_value1}\n{file2}: {cell_value2}"
                ws.cell(row=i+1, column=j+1).fill = green_fill
            else:
                ws.cell(row=i+1, column=j+1).value = f"{file1}: {cell_value1}\n{file2}: {cell_value2}"
                ws.cell(row=i+1, column=j+1).fill = red_fill

    # Save the result to File3
    wb.save(file3)

# Example usage
file1 = 'File1.csv'  # or 'File1.xlsx'
file2 = 'File2.xlsx'  # or 'File2.xlsx'
file3 = 'comparison_results.xlsx'  # result file (should be Excel to handle cell colors)

compare_files(file1, file2, file3)
